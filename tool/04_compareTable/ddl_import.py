#! python3

import openpyxl
import logging
import pprint
import os
from pathlib import Path
import sys
import json
import re

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.DEBUG)

def makeTableData(target_list):
    logging.debug('sheetnames: {}'.format(wb.sheetnames))
    for table_name in set(target_list) & set(wb.sheetnames): #ワークブックのシート名と一致する分だけ
        sheet = wb[table_name]
        scheme = table_name[:2]
        
        #連番用
        pre_table =''
        sno=1
        
        for cur_row in range(1, sheet.max_row + 1):
            table = sheet.cell(cur_row, 1).value
            col = sheet.cell(cur_row, 2).value
            type = sheet.cell(cur_row, 3).value
            len = sheet.cell(cur_row, 4).value
            
            if table != pre_table:
                sno = 1
            
            logging.debug('scheme: {}, table: {}, col: {}, type: {}, len: {}, sno: {}'.format(scheme, table, col, type, len, sno))
            
            all_tables[scheme].setdefault(table, {})
            all_tables[scheme][table].setdefault(col, {'type': type, 'len': len, 'sno':sno})
            
            pre_table = table
            sno += 1

def add_lib_ddl_all(filename, add_str):
    path_lo = path_to_lib / filename
    
    add_str_mod = 'import lib.' + add_str
    
    if path_lo.is_file():
        file_lo = open(str(path_lo), 'r+') #読み書きモード
        file_lo.seek(0) #readlines()は、現在の位置から読み取るので念の為offsecを0にしておく
        if add_str_mod not in [line.strip() for line in file_lo.readlines()]:
            file_lo.seek(0, os.SEEK_END)
            file_lo.write(add_str_mod + '\n')
    else:
        file_lo = open(str(path_lo), 'w') #書き込みモード
        file_lo.write(add_str_mod + '\n')

#start
logging.debug('program start ...')

#init
FILENAME = ''
target_list = []
path_to_excel = ''
path_to_lib = ''
all_tables = {}

#引数を取得
if len(sys.argv) > 1:
    json_str = sys.argv[1]
    data = json.loads(json_str)
    logging.info(f'受け取ったjson: {data}')
    
    #JSONをセット
    path_to_excel = Path(data['path_to_excel'])
    path_to_lib = Path(data['path_to_lib']) / 'ddl'
    FILENAME = data['filename']
    target_list = data['target']
    regex_str = data['regex']
    
else:
    logging.info(f'引数が渡されていません。')

#init
path_to_excel_files = path_to_excel.glob(FILENAME) #カレントのexcelフォルダを検索対象とする。

#main
for path in path_to_excel_files:
    logging.debug(f'path: {path}')
    for table in target_list:
        scheme = table[:2] #TARGEの先頭２文字は業務コードだと信じて
        all_tables[scheme] = {}
        all_tables[scheme]['name'] = path.stem

    wb = openpyxl.load_workbook(str(path))
    makeTableData(target_list)
    
    path_to_lib.mkdir(parents=True, exist_ok=True)
    
    path_to_lib.mkdir(parents=True, exist_ok=True)
    
    filename_regex = re.compile(fr'{regex_str}')
    mo = filename_regex.search(path.stem)
    if  mo == None:
        logging.error('ファイル名と正規表現が一致しません。正規表現の設定を見直してください。')
        sys.exit()
    logging.debug(f'regex_str: {regex_str}, mo: {mo.group(1)}')
    filename = mo.group(1) + '.py'
    
    file = open(path_to_lib / filename, 'w')
    file.write('all_tables =' + pprint.pformat(all_tables))
    file.close()
    
    #第1引数に追加するファイル名を、第2引数に追加する文字を指定する
    # add_lib_ddl_all('ddl_all.py', path.stem)

#end
logging.debug('program complete')