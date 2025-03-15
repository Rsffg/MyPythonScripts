#! python3

import openpyxl
import logging
import pprint
import os
from pathlib import Path
import sys
import json
import importlib
import re

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.DEBUG)

def makeTableData(table_list):
    logging.debug('sheetnames: {}'.format(wb.sheetnames))
    for table_name in table_list & set(wb.sheetnames): #ワークブックのシート名と一致する分だけ
        sheet = wb[table_name]
        scheme = table_name[:2]
                
        for cur_row in range(1, sheet.max_row + 1):
            str = sheet.cell(cur_row, 1).value
            split_str = str.split(",")
            table = split_str[0]
            all_datas[scheme].setdefault(table, {})
            next_sno = all_datas[scheme][table].get('sno',0) + 1
            all_datas[scheme][table]['sno'] = next_sno
            all_datas[scheme][table][next_sno] = {}
            
            for i in range(1, len(split_str)):
                logging.debug('cur_row:{}, i:{}, table:{}, next_sno:{}'.format(cur_row,i,table,next_sno))
                
                #lib_moduleにtableが無いとこけるのでチェック
                if lib_module.all_tables[scheme].get(table, '') == '':
                    logging.info(f'all_tables[{scheme}]に{table}は存在しないので、取込みできません')
                    continue
                
                #ループの前にリセットしておく
                tmp_column = ''
                tmp_sno = 0
                for tmp_column in lib_module.all_tables[scheme][table]:
                    tmp_sno = lib_module.all_tables[scheme][table][tmp_column]['sno']
                    logging.debug('table:{}, tmp_column:{}, tmp:sno:{}'.format(table,tmp_column,tmp_sno))
                    
                    if i == int(tmp_sno):
                        all_datas[scheme][table][next_sno][tmp_column] = split_str[i]
                        break #見つかったらループを抜け出す。

def import_mylib(name):
    if str(path_to_lib) not in sys.path:
        sys.path.insert(0, str(path_to_lib))

    return importlib.import_module(name)



#start    
logging.debug('program start')

#init
filename = ''
target_list = []
path_to_excel = ''
all_datas = {}

if len(sys.argv) > 1:
    json_str = sys.argv[1]
    data = json.loads(json_str)
    logging.debug(f'受け取ったjson: {data}')
    
    #セット
    path_to_excel = Path(data['path_to_excel'])
    path_to_lib = Path(data['path_to_lib'])
    filename = data['filename']
    table_list = set(data['target'])
    regex_str = data['regex']

path2excel = path_to_excel.glob(filename) #カレントのexcelフォルダを検索対象とする。

#main
for path in path2excel:
    for table in table_list:
        scheme = table[:2]
        all_datas[scheme] = {}
        all_datas[scheme]['name'] = path.stem #拡張子なしはstemを使う
    
    #excel読み込み
    wb = openpyxl.load_workbook(str(path))
    
    filename_regex = re.compile(fr'{regex_str}')
    
    mo = filename_regex.search(path.stem)
    logging.debug(f'regex_str: {regex_str}, target:{path.stem}')
    
    if mo == None:
        logging.info('正規表現で名前が取得できないので終了')
        sys.exit()
    logging.debug(f'mo: {mo.group(1)}')
    
    lib_name = mo.group(1)
    
    #module読み込み
    lib_module = import_mylib('ddl.' + lib_name)
    
    
    makeTableData(table_list)

    path_to_dml_lib = path_to_lib / 'dml'
    path_to_dml_lib.mkdir(parents=True, exist_ok=True)
    file = open(path_to_dml_lib / (lib_name + '.py'), 'w')
    file.write('all_datas =' + pprint.pformat(all_datas))
    file.close()
    logging.debug('all_datas: {}'.format(all_datas))
    


#end
logging.debug('program complete')