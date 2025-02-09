#! python3

import openpyxl
import logging
import pprint
import os

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')


def makeTableData(file_name, table_list):
    wb = openpyxl.load_workbook(target + '.xlsx')
    
    for table_name in table_list & set(wb.sheetnames): #ワークブックのシート名と一致する分だけ
        sheet = wb[table_name]
        scheme = table_name[:2]
                
        for cur_row in range(1, sheet.max_row + 1):
            table = sheet.cell(cur_row, 1).value
            col = sheet.cell(cur_row, 2).value
            type = sheet.cell(cur_row, 3).value
            len = sheet.cell(cur_row, 4).value
            
            logging.debug('scheme: {}, table: {}, col: {}, type: {}, len: {}'.format(scheme, table, col, type, len))
            
            all_tables[scheme].setdefault(table, {})
            all_tables[scheme][table].setdefault(col, {'type': type, 'len': len})

    file = open(file_name + '.py', 'w')
    file.write('all_tables =' + pprint.pformat(all_tables))
    file.close()
    
logging.debug('program start')
logging.debug(os.getcwd())

all_tables = {}
target='pkg' #自治体名
table_list = {'EB_TABLE','EA_TABLE','EK_TABLE','EE_TABLE'}

for table in table_list:
    scheme = table[:2]
    all_tables[scheme] = {}
    all_tables[scheme]['name'] = target

makeTableData(target, table_list)

logging.debug('program complete')